import streamlit as st
import pandas as pd
import pyodbc
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Database configuration
CONFIG = {
    'server': 'caappsdb,1435',
    'database': 'BCSSoft_ConAppSys',
    'username': 'Deepak',
    'password': 'Deepak@321',
    'driver': 'ODBC Driver 17 for SQL Server'
}

def get_connection():
    conn_str = f"DRIVER={{{CONFIG['driver']}}};SERVER={CONFIG['server']};DATABASE={CONFIG['database']};UID={CONFIG['username']};PWD={CONFIG['password']}"
    return pyodbc.connect(conn_str)

def create_excel(df: pd.DataFrame) -> BytesIO:
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Adjustments')
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    ws.freeze_panes = "A2"

    header_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    highlight_columns = ['CCAdj_Qty', 'OnHandQty']
    for col in ws.iter_cols():
        if col[0].value in highlight_columns:
            for cell in col:
                cell.font = Font(bold=True)

    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

    styled_output = BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

def get_data(start_date, end_date):
    sql_query = f"""
    WITH DetailWithAdjustment AS (
        SELECT 
            H.CounterCode,
            H.CCName,
            H.ConfirmedDt,
            D.Location,
            D.SKU,
            D.QtyCounted,
            D.SystemQty,
            D.QtyAdjusted AS CCDetail_QtyAdjusted,
            A.Qty AS CCAdj_Qty,
            A.AdjDt,
            A.SourceFrom,
            CASE 
                WHEN A.CounterCode IS NOT NULL THEN 'Yes' 
                ELSE 'No' 
            END AS HasAdjustment
        FROM [dbo].[tbCCHeader] H
        JOIN [dbo].[tbCCDetail] D ON H.CCHeaderID = D.CCHeaderID
        LEFT JOIN [dbo].[tbInvAdj] A 
            ON H.CounterCode = A.CounterCode 
            AND D.SKU = A.SKU
            AND A.SourceFrom = 'CCadj'
            AND A.AdjDt = EOMONTH(H.ConfirmedDt, -1)
        WHERE 
            H.CCTypeCode = 'FST' 
            AND H.ConfirmedDt >= '{start_date.strftime('%Y-%m-%d')}'
            AND H.ConfirmedDt <= '{end_date.strftime('%Y-%m-%d')}'
    )
    SELECT 
        d.CounterCode,
        d.CCName,
        d.SKU,
        d.ConfirmedDt,
        SUM(d.SystemQty) AS TotalSystemQty,
        SUM(d.QtyCounted) AS TotalQtyCounted,
        d.CCAdj_Qty,
        d.AdjDt AS AdjustmentDate,
        d.HasAdjustment
    FROM DetailWithAdjustment d
    WHERE d.HasAdjustment = 'Yes'
    GROUP BY 
        d.CounterCode,
        d.CCName,
        d.SKU,
        d.ConfirmedDt,
        d.CCDetail_QtyAdjusted,
        d.CCAdj_Qty,
        d.AdjDt,
        d.HasAdjustment
    """
    with get_connection() as conn:
        return pd.read_sql(sql_query, conn)

def main():
    st.title("Cycle Count Adjustments Report")

    today = datetime.today()
    default_start = today - timedelta(days=90)

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("From Date", default_start)
        refresh_clicked = st.button("🔄 Refresh Data")
    with col2:
        end_date = st.date_input("To Date", today)

    if start_date > end_date:
        st.error("❌ From Date must be before or equal to To Date.")
        return

    if 'df_cache' not in st.session_state or refresh_clicked:
        try:
            with st.spinner("📡 Loading data from database..."):
                df = get_data(start_date, end_date)
                df['OnHandQty'] = df['TotalSystemQty'] - df['CCAdj_Qty']
                df.drop(columns=['TotalSystemQty'], inplace=True)
                df = df[[
                    'CounterCode', 'CCName', 'SKU', 'ConfirmedDt',
                    'CCAdj_Qty', 'OnHandQty', 'TotalQtyCounted',
                    'AdjustmentDate', 'HasAdjustment'
                ]]
                st.session_state.df_cache = df
        except Exception as e:
            st.error(f"❌ Error loading data: {e}")
            return

    if 'df_cache' in st.session_state and not st.session_state.df_cache.empty:
        df = st.session_state.df_cache.copy()
        df_display = df.copy()
        df_display.index = range(1, len(df_display) + 1)
        st.success(f"✅ Loaded {len(df_display)} records.")
        st.dataframe(df_display, use_container_width=True)

        with col2:
            excel_data = create_excel(df)
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_data,
                file_name=f"CycleCount_Adjustments_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("⚠ No data available for the selected date range.")

if __name__ == "__main__":
    main()
