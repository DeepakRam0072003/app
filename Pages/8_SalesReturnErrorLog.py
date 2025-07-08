import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import pyodbc
from datetime import datetime, timedelta
import os
import io
import logging

# Database Configuration
bcs = {
    'server': 'caappsdb,1435',
    'database': 'BCSSoft_ConAppSys',
    'username': 'Deepak',
    'password': 'Deepak@321',
    'driver': 'ODBC Driver 17 for SQL Server',
    'timeout': 30
}

nav = {
    'server': 'nav18db',
    'database': 'EDLIVE',
    'username': 'barcode1',
    'password': 'barcode@1433',
    'driver': 'ODBC Driver 17 for SQL Server',
    'timeout': 30
}

# Logging setup
log_file = os.path.join(os.environ['USERPROFILE'], 'Documents', 'cr_report_scheduler.log')
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_bcs_connection():
    conn_str = f"DRIVER={{{bcs['driver']}}};SERVER={bcs['server']};DATABASE={bcs['database']};UID={bcs['username']};PWD={bcs['password']};Timeout={bcs['timeout']}"
    return pyodbc.connect(conn_str)

def get_nav_connection():
    conn_str = f"DRIVER={{{nav['driver']}}};SERVER={nav['server']};DATABASE={nav['database']};UID={nav['username']};PWD={nav['password']};Timeout={nav['timeout']}"
    return pyodbc.connect(conn_str)

def generate_cr_report_df(start_date, end_date):
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    sql = f"""
    SELECT 
        h.CustRtnHeaderId,
        h.CRNo,
        h.ClosedDt,
        d.CustRtnDetailId,
        d.CustRtnTypeCode,
        d.CreatedDt,
        log.LogTypeCode,
        log.LogStsCode,
        log.LogMsg
    FROM 
        [BCSSoft_ConAppSys].[dbo].[tbCustRtnHeader] h
    INNER JOIN 
        [BCSSoft_ConAppSys].[dbo].[tbCustRtnDetail] d 
        ON h.CustRtnHeaderId = d.CustRtnHeaderId
    INNER JOIN 
        [BCSSoft_ConAppSys].[dbo].[tbIntgNavLog] log 
        ON log.RefHdrId = d.CustRtnHeaderId 
        AND log.RefDtlId = d.CustRtnDetailId
    WHERE 
        d.CustRtnTypeCode IN ('SalesReconcile', 'Refund', 'Exchange')
        AND log.LogStsCode = 'E'
        AND d.CreatedDt >= '{start_date_str}'
        AND d.CreatedDt <= '{end_date_str}'
    ORDER BY 
        d.CreatedDt DESC
    """

    with get_bcs_connection() as conn:
        df = pd.read_sql(sql, conn)

    if df.empty:
        return pd.DataFrame()

    def fix_crno(crno):
        if pd.isna(crno):
            return None
        try:
            if not isinstance(crno, str) or not crno.startswith("CR_"):
                return None
            part = crno.split('CR_')[-1].replace('_', '')
            if len(part) > 10 and part[8:10] == '20':
                part = part[:8] + part[10:]
            return part if part else None
        except Exception:
            return None

    df['Document No_'] = df['CRNo'].apply(fix_crno)
    doc_nos = [doc for doc in df['Document No_'].unique().tolist() if doc is not None]

    def chunks(lst, n):
        for i in range(0, len(lst), n):
            yield lst[i:i + n]

    nav_docs_set = set()
    with get_nav_connection() as nav_conn:
        for batch in chunks(doc_nos, 1000):
            in_clause = ",".join(f"'{doc}'" for doc in batch if doc)
            if not in_clause:
                continue
                
            nav_sql = f"""
            SELECT DISTINCT [Document No_]
            FROM [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Item Ledger Entry]
            WHERE [Entry Type] = '1'
              AND [Document No_] IN ({in_clause})
            """
            try:
                nav_df = pd.read_sql(nav_sql, nav_conn)
                nav_docs_set.update(nav_df['Document No_'].tolist())
            except Exception:
                continue

    df['NAV Status'] = df['Document No_'].apply(
        lambda x: 'Posted' if x and x in nav_docs_set else 'Not Posted'
    )

    # Columns to hide
    columns_to_hide = ['CustRtnHeaderId', 'CustRtnDetailId', 'LogTypeCode']
    
    # Reorder and select columns to display with LogMsg first
    cols = ['LogMsg', 'CRNo', 'Document No_', 'NAV Status'] + \
           [col for col in df.columns if col not in columns_to_hide + ['LogMsg', 'CRNo', 'Document No_', 'NAV Status']]
    
    # Remove duplicates in case we moved them
    cols = list(dict.fromkeys(cols))
    
    return df[cols]

def df_to_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "CA Sales Return Report (NAV)"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if df.columns[c_idx - 1] == 'NAV Status':
                cell.font = Font(color="008000" if value == 'Posted' else "FF0000")

    for col_cells in ws.columns:
        max_length = 0
        column_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column_letter].width = max_length + 2

    with io.BytesIO() as output:
        wb.save(output)
        data = output.getvalue()
    return data

def main():
    st.set_page_config(layout="wide")
    st.title("CA Sales Return Error Report (NAV)")

    # Date range selection (default last 3 months)
    today = datetime.today()
    default_start = today - timedelta(days=90)
    
    # Create two columns for date inputs
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input("From Date", default_start)
        refresh_clicked = st.button("Refresh Data")
    
    with col2:
        end_date = st.date_input("To Date", today)
        if 'df_report' in st.session_state and not st.session_state.df_report.empty:
            # Filter only "Not Posted" records for download
            not_posted_df = st.session_state.df_report[
                st.session_state.df_report['NAV Status'] == 'Not Posted'
            ]
            if not not_posted_df.empty:
                excel_bytes = df_to_excel_bytes(not_posted_df)
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_bytes,
                    file_name=f"CA_Sales_Return_Report_NAV_Not_Posted_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='download_excel'
                )

    if start_date > end_date:
        st.error("From Date must be before or equal to To Date.")
        return

    # Load data when app starts or when refresh is clicked
    if refresh_clicked or 'df_report' not in st.session_state:
        with st.spinner("Fetching data..."):
            try:
                df_report = generate_cr_report_df(start_date, end_date)
                st.session_state.df_report = df_report
            except Exception as e:
                st.error(f"Error: {str(e)}")
                return

    if 'df_report' in st.session_state:
        df_report = st.session_state.df_report
        if df_report.empty:
            st.warning("No data found for the selected date range with error status.")
        else:
            # Filter to show only "Not Posted" records
            not_posted_df = df_report[df_report['NAV Status'] == 'Not Posted']
            
            if not not_posted_df.empty:
                st.success(f"Found {len(not_posted_df)} records with 'Not Posted' status.")
                
                # Display with index starting at 1 and LogMsg first
                display_df = not_posted_df.copy()
                display_df.index = display_df.index + 1
                st.dataframe(display_df, use_container_width=True)
            else:
                st.warning("No records with 'Not Posted' status found for the selected date range.")

if __name__ == '__main__':
    main()