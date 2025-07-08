import pyodbc
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import streamlit as st
from ws_utils import inject_websocket_code
from ws_triggers import ws_trigger

# SQL connection string
edlive_conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=nav18db;'
    'DATABASE=EDLIVE;'
    'UID=barcode1;'
    'PWD=barcode@1433;'
    'Trusted_Connection=no;'
)

# SQL query
sql_query = """ 
WITH BaseData AS (
    SELECT 
        h.No_ AS [Transfer No],
        h.[Transfer-from Code],
        h.[Transfer-to Code],
        h.[Created by WS],
        'Open' AS Status,
        h.[Posting Date],
        h.[External Document No_],
        h.[External Document No_ 2],
        l.[Document No_] AS [Transfer Document No],
        l.[Outstanding Quantity] AS [Transfer Quantity],
        re.Quantity AS [Reserved Quantity],
        l.[Item No_],
        l.[Line No_],
        ISNULL(ile.RemainingQuantity, 0) AS [Initial Stock],
        re.[Source ID],
        re.[Source Ref_ No_],
        CASE 
            WHEN re.Positive = 0 THEN 'From ILE (Available Stock)'
            WHEN re.Positive = 1 THEN 'From Pending Transfer Order'
            ELSE 'YetReserved'
        END AS [Reservation Source]
    FROM 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Header] h
    INNER JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Transfer Line] l
        ON h.No_ = l.[Document No_]
    LEFT JOIN (
        SELECT 
            [Item No_],  
            [Location Code],  
            SUM([Remaining Quantity]) AS RemainingQuantity 
        FROM 
            [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Item Ledger Entry]
        WHERE 
            [Remaining Quantity] > 0
            AND [Entry Type] in('4','2','1','0') 
        GROUP BY 
            [Item No_], 
            [Location Code]
    ) ile
        ON l.[Item No_] = ile.[Item No_]
        AND h.[Transfer-from Code] = ile.[Location Code]
    LEFT JOIN 
        [EDLIVE].[dbo].[Eastern Decorator Sdn_ Bhd_$Reservation Entry] re
        ON h.No_ = re.[Source ID]
        AND l.[Line No_] = re.[Source Ref_ No_]
        AND re.[Source ID] IS NOT NULL
        AND LTRIM(RTRIM(re.[Source ID])) <> ''
    WHERE 
        h.[Created By User ID] = 'EADECO\\CAAPI' 
        AND h.Status = 0
        AND (
            h.[External Document No_ 2] IS NULL 
            OR h.[External Document No_ 2] NOT LIKE 'E\\_%' ESCAPE '\\'
        )
)

SELECT 
    [Transfer No],
    [Transfer-from Code],
    [Transfer-to Code],
    [Created by WS],
    Status,
    [Posting Date],
    [External Document No_],
    [External Document No_ 2],
    [Transfer Document No],
    [Item No_],
    [Line No_],
    [Transfer Quantity],
    [Reserved Quantity],
    CASE 
        WHEN SUM(
            CASE 
                WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                ELSE ISNULL([Reserved Quantity], 0) 
            END
        ) OVER (
            PARTITION BY [Transfer Document No], [Item No_], [Line No_]
            ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) IS NOT NULL
        THEN 
            [Transfer Quantity] - 
            SUM(
                CASE 
                    WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                    ELSE ISNULL([Reserved Quantity], 0) 
                END
            ) OVER (
                PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )
        ELSE [Transfer Quantity]
    END AS [Unreserved Quantity],
    [Initial Stock],
    [Initial Stock] + 
        SUM(ISNULL([Reserved Quantity], 0)) OVER (
            PARTITION BY [Item No_], [Transfer-from Code]
            ORDER BY [Posting Date], [Transfer Document No], [Line No_]
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS [Available Quantity After Reservation],
    CASE
        WHEN 
            ([Initial Stock] + 
            SUM(ISNULL([Reserved Quantity], 0)) OVER (
                PARTITION BY [Item No_], [Transfer-from Code]
                ORDER BY [Posting Date], [Transfer Document No], [Line No_]
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            )) = 0
        THEN 
            CASE 
                WHEN SUM(
                    CASE 
                        WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                        ELSE ISNULL([Reserved Quantity], 0) 
                    END
                ) OVER (
                    PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                    ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                    ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ) IS NOT NULL
                THEN 
                    [Transfer Quantity] -
                    SUM(
                        CASE 
                            WHEN [Reserved Quantity] < 0 THEN ABS([Reserved Quantity]) 
                            ELSE ISNULL([Reserved Quantity], 0) 
                        END
                    ) OVER (
                        PARTITION BY [Transfer Document No], [Item No_], [Line No_]
                        ORDER BY [Posting Date], [Source ID], [Source Ref_ No_]
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    )
                ELSE [Transfer Quantity]
            END
        ELSE 0
    END AS [BalanceToReserved],
    [Source ID],
    [Source Ref_ No_],
    [Reservation Source]
FROM BaseData
ORDER BY 
    [Posting Date] DESC,
    [Item No_],
    [Line No_];
"""

def get_data(from_date=None, to_date=None):
    try:
        conn = pyodbc.connect(edlive_conn_str)
        df = pd.read_sql(sql_query, conn)
        
        if df is not None and 'Posting Date' in df.columns:
            df['Posting Date'] = pd.to_datetime(df['Posting Date'])
            if from_date and to_date:
                mask = (df['Posting Date'] >= pd.to_datetime(from_date)) & (df['Posting Date'] <= pd.to_datetime(to_date))
                df = df.loc[mask]
        
        ws_trigger.send_notification(
            "Data refreshed successfully!",
            channel="inventory",
            category="success"
        )
        
        return df
    except Exception as e:
        error_msg = f"Database error: {str(e)}"
        ws_trigger.send_notification(
            error_msg,
            channel="inventory",
            category="error"
        )
        st.error(f"❌ {error_msg}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()

def format_excel_report(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'B2'
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"❌ Excel formatting error: {str(e)}")
        return False

def generate_excel_report(df):
    try:
        output_path = r"Z:\\CavsNavErrors\\TOYetReservedNoInventory"
        os.makedirs(output_path, exist_ok=True)

        filename = f"No_Inventory_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file = os.path.join(output_path, filename)

        export_df = df[
            (df['Reservation Source'] == 'YetReserved') & 
            (df['BalanceToReserved'] > 0)
        ].copy()

        export_df['ILEPostingdate'] = pd.to_datetime(export_df['Posting Date']).dt.to_period('M').dt.to_timestamp().dt.strftime('%Y-%m-%d')

        grouped = export_df.groupby(
            ['ILEPostingdate', 'Item No_', 'Transfer-from Code', 'External Document No_', 'External Document No_ 2'],
            as_index=False
        )['BalanceToReserved'].sum()

        grouped = grouped.rename(columns={
            'Transfer-from Code': 'Location Code'
        })

        grouped = grouped[[ 'ILEPostingdate', 'Item No_', 'Location Code', 'External Document No_', 'External Document No_ 2', 'BalanceToReserved' ]]

        grouped.to_excel(output_file, index=False)

        if format_excel_report(output_file):
            return output_file
        return None
    except Exception as e:
        st.error(f"❌ Report generation error: {str(e)}")
        return None

def main():
    st.set_page_config(layout="wide")
    st.title("📋 No Inventory Report")
    
    # Inject WebSocket client
    st.markdown(inject_websocket_code(), unsafe_allow_html=True)
    
    # Add WebSocket status indicator
    st.sidebar.markdown("""
    <div id="ws-status" style="padding: 8px; border-radius: 4px; background: #f0f0f0; margin-top: 20px;">
        <small>Connection: <span style="color: gray">Connecting...</span></small>
    </div>
    """, unsafe_allow_html=True)

    # Initialize session state
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'report_path' not in st.session_state:
        st.session_state.report_path = None
    if 'filtered_df' not in st.session_state:
        st.session_state.filtered_df = None
    if 'last_update' not in st.session_state:
        st.session_state.last_update = datetime.now()

    today = datetime.now().date()
    default_from_date = today - timedelta(days=90)

    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", value=default_from_date)
    with col2:
        to_date = st.date_input("To Date", value=today)

    if st.session_state.df is None or 'last_from_date' not in st.session_state or \
       st.session_state.last_from_date != from_date or st.session_state.last_to_date != to_date:
        
        with st.spinner("Fetching data..."):
            st.session_state.df = get_data(from_date=from_date, to_date=to_date)
            st.session_state.last_from_date = from_date
            st.session_state.last_to_date = to_date
            
            if st.session_state.df is not None:
                filtered_df = st.session_state.df[
                    (st.session_state.df['Reservation Source'] == 'YetReserved') &
                    (st.session_state.df['BalanceToReserved'] > 0)
                ].copy()

                filtered_df['Posting Date'] = pd.to_datetime(filtered_df['Posting Date'])
                grouped_df = filtered_df.groupby(
                    ['Transfer No', 'Item No_', 'Transfer-from Code', 'Transfer-to Code'],
                    as_index=False
                ).agg({
                    'Posting Date': 'min',
                    'External Document No_': 'first',
                    'External Document No_ 2': 'first',
                    'Transfer Quantity': 'sum',
                    'BalanceToReserved': 'sum'
                })

                grouped_df['Posting Date'] = grouped_df['Posting Date'].dt.strftime('%Y-%m-%d')
                st.session_state.filtered_df = grouped_df

    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔄 Refresh Data"):
            with st.spinner("Refreshing data..."):
                st.session_state.df = get_data(from_date=from_date, to_date=to_date)
                if st.session_state.df is not None:
                    filtered_df = st.session_state.df[
                        (st.session_state.df['Reservation Source'] == 'YetReserved') &
                        (st.session_state.df['BalanceToReserved'] > 0)
                    ].copy()

                    filtered_df['Posting Date'] = pd.to_datetime(filtered_df['Posting Date'])
                    grouped_df = filtered_df.groupby(
                        ['Transfer No', 'Item No_', 'Transfer-from Code', 'Transfer-to Code'],
                        as_index=False
                    ).agg({
                        'Posting Date': 'min',
                        'External Document No_': 'first',
                        'External Document No_ 2': 'first',
                        'Transfer Quantity': 'sum',
                        'BalanceToReserved': 'sum'
                    })

                    grouped_df['Posting Date'] = grouped_df['Posting Date'].dt.strftime('%Y-%m-%d')
                    st.session_state.filtered_df = grouped_df

    with col2:
        if st.button("📤 Generate Excel Report"):
            if st.session_state.df is None:
                st.warning("⚠️ No data available to generate report")
            else:
                with st.spinner("Generating Excel report..."):
                    path = generate_excel_report(st.session_state.df)
                    if path:
                        st.session_state.report_path = path
                        ws_trigger.send_notification(
                            "Excel report generated!",
                            channel="inventory",
                            category="success"
                        )
                        st.success("✅ Excel report generated!")

                        with open(path, "rb") as f:
                            st.download_button(
                                label="📥 Download Excel Report",
                                data=f,
                                file_name=os.path.basename(path),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

    if st.session_state.filtered_df is not None:
        df_display = st.session_state.filtered_df.copy()
        df_display.index = df_display.index + 1  # Start from 1
        df_display.index.name = "Row"
        
        st.markdown("""
        <div id="inventory-table-container">
            <!-- Table will be updated via WebSocket -->
        </div>
        <script>
            document.addEventListener('ws-data', (event) => {
                if (event.detail.type === 'inventory_update') {
                    console.log('Received inventory update:', event.detail.data);
                }
            });
        </script>
        """, unsafe_allow_html=True)
        
        st.dataframe(df_display, use_container_width=True)

if __name__ == "__main__":
    main()